from __future__ import annotations

import argparse
import datetime as dt
import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from utils import normalize_code, normalize_text, safe_int, sqlite_connect

DEPT_KEYWORDS = ('departamento', 'partido', 'comuna')
POP_KEYWORDS = ('total de población', 'población total', 'total poblacion', 'poblacion total')


def norm(v) -> str:
    return normalize_text(v) or ''


def find_header(ws):
    for r in range(1, min(ws.max_row, 80) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        nvalues = [norm(v) for v in values]
        dept_cols = [i + 1 for i, v in enumerate(nvalues) if any(k in v for k in DEPT_KEYWORDS)]
        pop_cols = [i + 1 for i, v in enumerate(nvalues) if any(norm(k) in v for k in POP_KEYWORDS)]
        if dept_cols and pop_cols:
            return r, dept_cols[0], pop_cols[0]
    return None


def parse_file(path: Path, province_hint: str | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    rows: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        found = find_header(ws)
        if not found:
            continue
        header_row, dept_col, pop_col = found
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, dept_col).value
            pop = ws.cell(r, pop_col).value
            if name is None:
                continue
            name_s = str(name).strip()
            if not name_s or normalize_text(name_s) in {'total', 'total provincia', 'total del pais'}:
                continue
            pop_i = safe_int(pop)
            if pop_i is None:
                continue
            rows.append({
                'archivo': path.name,
                'sheet': sheet,
                'provincia_hint': province_hint,
                'nombre_departamento': name_s,
                'nombre_normalizado': normalize_text(name_s),
                'poblacion_total': pop_i,
            })
    return rows


def infer_province_code_from_filename(path: Path) -> str | None:
    m = re.search(r'c2022_(\d{2})_', path.name)
    return m.group(1) if m else None


def main() -> None:
    parser = argparse.ArgumentParser(description='Importa población Censo 2022 por departamento/partido/comuna desde XLSX INDEC descargados.')
    parser.add_argument('--input-dir', required=True)
    parser.add_argument('--sqlite', required=True)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    files = sorted(input_dir.glob('*.xlsx'))
    if not files:
        raise SystemExit(f'No hay XLSX en {input_dir}')

    all_rows = []
    for path in files:
        prov = infer_province_code_from_filename(path)
        rows = parse_file(path, prov)
        print(f'{path.name}: {len(rows)} filas detectadas')
        all_rows.extend(rows)

    df = pd.DataFrame(all_rows)
    if df.empty:
        raise SystemExit('No se detectaron filas departamentales. Revisar formato de Excel.')

    if args.dry_run:
        print(df.head(20).to_string(index=False))
        return

    conn = sqlite_connect(args.sqlite)
    now = dt.datetime.now(dt.UTC).isoformat()
    conn.execute(
        """INSERT OR IGNORE INTO fuente_catalogo
           (id_fuente,nombre_fuente,organismo,url_origen,tipo_fuente,fecha_extraccion,observaciones)
           VALUES (?,?,?,?,?,?,?)""",
        ('indec_censo_2022_departamentos', 'Censo 2022 por departamento/partido/comuna', 'INDEC', str(input_dir), 'oficial', now, 'Importado desde XLSX INDEC mediante heurística de encabezados.')
    )

    for _, row in df.iterrows():
        prov_code = normalize_code(row.get('provincia_hint'), 2)
        # Sin código oficial en el Excel parseado, se genera ID provisional por provincia + nombre normalizado.
        # Luego se debe reconciliar contra Georef por código/nombre con build_full_gpkg o una revisión manual.
        dept_norm = row['nombre_normalizado']
        dept_id = f'departamento_match:{prov_code or "xx"}:{dept_norm}'
        conn.execute(
            """INSERT OR IGNORE INTO geo_entidad
               (id_entidad,tipo_entidad,codigo_indec,nombre,nombre_normalizado,provincia_id,
                fuente_geografica,clasificacion_fuente,metodo_geometria,nivel_confianza,anio_fuente,atributos_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (dept_id, 'departamento_partido_comuna', None, row['nombre_departamento'], dept_norm,
             f'provincia:{prov_code}' if prov_code else None, 'indec_censo_2022_departamentos',
             'oficial_procesada', 'excel_indec_match_pendiente_georef', 'media', 2022,
             '{"archivo":"%s","sheet":"%s"}' % (row['archivo'], row['sheet']))
        )
        conn.execute(
            """INSERT OR REPLACE INTO censo_poblacion
               (id_entidad,anio_censo,poblacion_total,fuente_censo,clasificacion_fuente,metodo_dato,nivel_confianza,fecha_extraccion,observaciones)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (dept_id, 2022, int(row['poblacion_total']), 'indec_censo_2022_departamentos',
             'oficial_procesada', 'xlsx_indec_departamental', 'media', now, 'Requiere reconciliación final con código Georef/INDEC.')
        )
    conn.commit()
    conn.close()
    print(f'Importadas {len(df):,} filas departamentales a {args.sqlite}')


if __name__ == '__main__':
    main()
