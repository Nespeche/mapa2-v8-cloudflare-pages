from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from utils import normalize_code, normalize_text, safe_float, safe_int, sqlite_connect

PROV_NAME_TO_CODE = {
    normalize_text('Ciudad Autónoma de Buenos Aires'): '02',
    normalize_text('Buenos Aires'): '06',
    normalize_text('Catamarca'): '10',
    normalize_text('Chaco'): '22',
    normalize_text('Chubut'): '26',
    normalize_text('Córdoba'): '14',
    normalize_text('Corrientes'): '18',
    normalize_text('Entre Ríos'): '30',
    normalize_text('Formosa'): '34',
    normalize_text('Jujuy'): '38',
    normalize_text('La Pampa'): '42',
    normalize_text('La Rioja'): '46',
    normalize_text('Mendoza'): '50',
    normalize_text('Misiones'): '54',
    normalize_text('Neuquén'): '58',
    normalize_text('Río Negro'): '62',
    normalize_text('Salta'): '66',
    normalize_text('San Juan'): '70',
    normalize_text('San Luis'): '74',
    normalize_text('Santa Cruz'): '78',
    normalize_text('Santa Fe'): '82',
    normalize_text('Santiago del Estero'): '86',
    normalize_text('Tierra del Fuego, Antártida e Islas del Atlántico Sur'): '94',
    normalize_text('Tucumán'): '90',
}


def init_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        '''
        DROP TABLE IF EXISTS fuente_catalogo;
        DROP TABLE IF EXISTS geo_entidad;
        DROP TABLE IF EXISTS censo_poblacion;
        DROP TABLE IF EXISTS refeglo_gobiernos_locales;
        DROP TABLE IF EXISTS geo_relacion;
        DROP TABLE IF EXISTS staging_match_log;
        DROP TABLE IF EXISTS diagnostico_cobertura;
        DROP TABLE IF EXISTS diagnostico_faltantes;
        DROP TABLE IF EXISTS control_cobertura;

        CREATE TABLE fuente_catalogo (
            id_fuente TEXT PRIMARY KEY,
            nombre_fuente TEXT NOT NULL,
            organismo TEXT,
            url_origen TEXT,
            tipo_fuente TEXT NOT NULL,
            licencia TEXT,
            fecha_publicacion TEXT,
            fecha_actualizacion TEXT,
            fecha_extraccion TEXT,
            observaciones TEXT
        );

        CREATE TABLE geo_entidad (
            id_entidad TEXT PRIMARY KEY,
            tipo_entidad TEXT NOT NULL,
            codigo_indec TEXT,
            codigo_georef TEXT,
            codigo_refeglo TEXT,
            nombre TEXT NOT NULL,
            nombre_normalizado TEXT,
            provincia_id TEXT,
            departamento_id TEXT,
            municipio_id TEXT,
            gobierno_local_id TEXT,
            localidad_id TEXT,
            fuente_geografica TEXT,
            clasificacion_fuente TEXT NOT NULL,
            metodo_geometria TEXT NOT NULL,
            nivel_confianza TEXT NOT NULL,
            anio_fuente INTEGER,
            centroid_lat REAL,
            centroid_lon REAL,
            geom_wkt TEXT,
            atributos_json TEXT
        );

        CREATE TABLE censo_poblacion (
            id_poblacion INTEGER PRIMARY KEY AUTOINCREMENT,
            id_entidad TEXT NOT NULL,
            anio_censo INTEGER NOT NULL,
            poblacion_total INTEGER,
            poblacion_varones INTEGER,
            poblacion_mujeres INTEGER,
            poblacion_viviendas_particulares INTEGER,
            poblacion_viviendas_colectivas INTEGER,
            poblacion_situacion_calle INTEGER,
            viviendas_total INTEGER,
            viviendas_particulares INTEGER,
            viviendas_colectivas INTEGER,
            superficie_km2 REAL,
            densidad_hab_km2 REAL,
            fuente_censo TEXT,
            clasificacion_fuente TEXT NOT NULL,
            metodo_dato TEXT NOT NULL,
            nivel_confianza TEXT NOT NULL,
            fecha_extraccion TEXT,
            observaciones TEXT
        );

        CREATE TABLE refeglo_gobiernos_locales (
            codigo_refeglo TEXT PRIMARY KEY,
            provincia TEXT,
            departamento TEXT,
            nombre_gobierno_local TEXT,
            categoria_gobierno TEXT,
            poblacion_censo_2010 INTEGER,
            localidades_administradas TEXT,
            latitud REAL,
            longitud REAL,
            sitio_web TEXT,
            fuente TEXT
        );

        CREATE TABLE geo_relacion (
            id_relacion INTEGER PRIMARY KEY AUTOINCREMENT,
            id_origen TEXT NOT NULL,
            id_destino TEXT NOT NULL,
            tipo_relacion TEXT NOT NULL,
            porcentaje_area REAL,
            porcentaje_poblacion REAL,
            metodo TEXT NOT NULL,
            fuente TEXT,
            atributos_json TEXT DEFAULT '{}',
            UNIQUE(id_origen, id_destino, tipo_relacion, metodo)
        );

        CREATE TABLE staging_match_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entidad_tipo TEXT,
            fuente TEXT,
            codigo_fuente TEXT,
            nombre_fuente TEXT,
            id_entidad_match TEXT,
            score REAL,
            metodo_match TEXT,
            estado TEXT,
            detalle TEXT,
            created_at TEXT
        );

        CREATE TABLE diagnostico_cobertura (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ejecutado_en TEXT NOT NULL,
            tipo_entidad TEXT NOT NULL,
            entidades INTEGER NOT NULL,
            con_codigo_indec INTEGER NOT NULL,
            con_codigo_georef INTEGER NOT NULL,
            con_codigo_refeglo INTEGER NOT NULL,
            con_centroide INTEGER NOT NULL,
            con_geom_wkt INTEGER NOT NULL,
            con_poblacion_2022 INTEGER NOT NULL,
            poblacion_oficial_directa INTEGER NOT NULL,
            poblacion_oficial_procesada INTEGER NOT NULL,
            poblacion_estimada INTEGER NOT NULL,
            relaciones_padre INTEGER NOT NULL,
            relaciones_huerfanas INTEGER NOT NULL,
            estado TEXT NOT NULL,
            observaciones TEXT
        );

        CREATE TABLE diagnostico_faltantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ejecutado_en TEXT NOT NULL,
            severidad TEXT NOT NULL,
            categoria TEXT NOT NULL,
            entidad_tipo TEXT,
            id_entidad TEXT,
            descripcion TEXT NOT NULL,
            accion_recomendada TEXT,
            fuente_sugerida TEXT
        );

        CREATE TABLE control_cobertura (
            tipo_entidad TEXT PRIMARY KEY,
            entidades INTEGER,
            con_centroides INTEGER,
            con_geom_wkt INTEGER,
            con_poblacion_2022 INTEGER,
            observaciones TEXT
        );
        '''
    )


def insert_sources(conn: sqlite3.Connection) -> None:
    now = dt.datetime.utcnow().isoformat()
    rows = [
        ('georef', 'Georef Argentina - base completa', 'Datos Argentina / IGN / INDEC / BAHRA', 'https://apis.datos.gob.ar/georef/api/v2.0', 'oficial', 'Datos Argentina', None, None, now, 'Descarga normalizada de unidades territoriales.'),
        ('indec_censo_2022_resumen', 'Censo 2022 resumen por jurisdicción', 'INDEC', 'https://www.indec.gob.ar/ftp/cuadros/poblacion/c2022_tp_c_resumen.xlsx', 'oficial', None, None, None, now, 'Población total por jurisdicción.'),
        ('indec_censo_2022_gobiernos_locales', 'Censo 2022 gobiernos locales', 'INDEC', 'https://www.indec.gob.ar/ftp/cuadros/poblacion/c2022_tp_gobierno_local_c1.xlsx', 'oficial', None, None, None, now, 'Viviendas y población por gobierno local.'),
        ('refeglo', 'Registro Federal de Gobiernos Locales', 'Ministerio del Interior', 'https://infra.datos.gob.ar/catalog/otros/dataset/14/distribution/14.1/download/datos-refeglo_25-09-2023.csv', 'oficial', None, None, '2023-09-25', now, 'Registro oficial de gobiernos locales con población 2010 y localidades administradas.'),
    ]
    conn.executemany('INSERT INTO fuente_catalogo VALUES (?,?,?,?,?,?,?,?,?,?)', rows)


def load_georef_provinces(seed_dir: Path, conn: sqlite3.Connection) -> None:
    path = seed_dir / 'provincias_georef_centroides.csv'
    if not path.exists():
        return
    df = pd.read_csv(path, dtype=str)
    for _, row in df.iterrows():
        code = normalize_code(row.get('id'), 2)
        if not code:
            continue
        name = str(row.get('nombre') or row.get('nombre_completo') or '').strip()
        lat = safe_float(row.get('centroide_lat'))
        lon = safe_float(row.get('centroide_lon'))
        conn.execute(
            '''INSERT OR REPLACE INTO geo_entidad
               (id_entidad,tipo_entidad,codigo_indec,codigo_georef,nombre,nombre_normalizado,
                fuente_geografica,clasificacion_fuente,metodo_geometria,nivel_confianza,anio_fuente,
                centroid_lat,centroid_lon,atributos_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (f'provincia:{code}', 'provincia', code, code, name, normalize_text(name),
             'georef', 'oficial_directa', 'centroide_semilla_georef', 'alta', 2026,
             lat, lon, '{}')
        )


def load_censo_resumen(seed_dir: Path, conn: sqlite3.Connection) -> None:
    path = seed_dir / 'c2022_tp_c_resumen.xlsx'
    if not path.exists():
        return
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    now = dt.datetime.utcnow().isoformat()
    for r in range(5, ws.max_row + 1):
        jurisd = ws.cell(r, 1).value
        if not jurisd:
            continue
        name_norm = normalize_text(jurisd)
        if name_norm == normalize_text('Total del país'):
            entity_id = 'pais:argentina'
            conn.execute(
                '''INSERT OR IGNORE INTO geo_entidad
                   (id_entidad,tipo_entidad,codigo_indec,codigo_georef,nombre,nombre_normalizado,
                    fuente_geografica,clasificacion_fuente,metodo_geometria,nivel_confianza,anio_fuente,atributos_json)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
                (entity_id, 'pais', 'ARG', 'ARG', 'Argentina', 'argentina', 'indec_censo_2022_resumen',
                 'oficial_directa', 'sin_geometria_semilla', 'alta', 2022, '{}')
            )
        else:
            code = PROV_NAME_TO_CODE.get(name_norm)
            if not code:
                continue
            entity_id = f'provincia:{code}'
        conn.execute(
            '''INSERT INTO censo_poblacion
               (id_entidad,anio_censo,poblacion_total,poblacion_varones,poblacion_mujeres,
                poblacion_viviendas_particulares,poblacion_viviendas_colectivas,poblacion_situacion_calle,
                fuente_censo,clasificacion_fuente,metodo_dato,nivel_confianza,fecha_extraccion)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (entity_id, 2022,
             safe_int(ws.cell(r, 2).value), safe_int(ws.cell(r, 6).value), safe_int(ws.cell(r, 10).value),
             safe_int(ws.cell(r, 3).value), safe_int(ws.cell(r, 4).value), safe_int(ws.cell(r, 5).value),
             'indec_censo_2022_resumen', 'oficial_directa', 'cuadro_resumen_indec', 'alta', now)
        )


def load_censo_gobiernos(seed_dir: Path, conn: sqlite3.Connection) -> None:
    path = seed_dir / 'c2022_tp_gobierno_local_c1.xlsx'
    if not path.exists():
        return
    wb = load_workbook(path, data_only=True)
    ws = wb['Total del país']
    now = dt.datetime.utcnow().isoformat()
    for r in range(6, ws.max_row + 1):
        prov_code = normalize_code(ws.cell(r, 1).value, 2)
        prov_name = ws.cell(r, 2).value
        gl_code = normalize_code(ws.cell(r, 3).value, 6)
        categoria = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if not gl_code or not name or not prov_code:
            continue
        entity_id = f'gobierno_local:{gl_code}'
        conn.execute(
            '''INSERT OR IGNORE INTO geo_entidad
               (id_entidad,tipo_entidad,codigo_indec,codigo_refeglo,nombre,nombre_normalizado,provincia_id,
                fuente_geografica,clasificacion_fuente,metodo_geometria,nivel_confianza,anio_fuente,atributos_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (entity_id, 'gobierno_local', gl_code, gl_code, str(name).strip(), normalize_text(name), f'provincia:{prov_code}',
             'indec_censo_2022_gobiernos_locales', 'oficial_directa', 'sin_geometria_semilla', 'alta', 2022,
             '{"categoria": "%s", "provincia": "%s"}' % (str(categoria or '').replace('"','\\"'), str(prov_name or '').replace('"','\\"')))
        )
        conn.execute(
            '''INSERT INTO censo_poblacion
               (id_entidad,anio_censo,poblacion_total,viviendas_total,viviendas_particulares,
                poblacion_viviendas_particulares,viviendas_colectivas,
                fuente_censo,clasificacion_fuente,metodo_dato,nivel_confianza,fecha_extraccion)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (entity_id, 2022,
             safe_int(ws.cell(r, 7).value), safe_int(ws.cell(r, 6).value), safe_int(ws.cell(r, 8).value),
             safe_int(ws.cell(r, 9).value), safe_int(ws.cell(r, 10).value),
             'indec_censo_2022_gobiernos_locales', 'oficial_directa', 'cuadro_gobierno_local_indec', 'alta', now)
        )


def load_refeglo(seed_dir: Path, conn: sqlite3.Connection) -> None:
    path = seed_dir / 'datos-refeglo_25-09-2023.csv'
    if not path.exists():
        return
    df = pd.read_csv(path, dtype=str)
    for _, row in df.iterrows():
        code = normalize_code(row.get('cod_gl_res144_indec'), 6)
        if not code:
            continue
        lat = safe_float(row.get('latitud'))
        lon = safe_float(row.get('longitud'))
        # En el CSV observado, algunas filas traen lat/lon invertidas. Corregimos si la latitud cae fuera del rango argentino plausible.
        if lat is not None and lon is not None and abs(lat) > 55 and abs(lon) <= 55:
            lat, lon = lon, lat
        pobl2010 = safe_int(row.get('poblacion_censo_2010'))
        conn.execute(
            '''INSERT OR REPLACE INTO refeglo_gobiernos_locales
               (codigo_refeglo,provincia,departamento,nombre_gobierno_local,categoria_gobierno,
                poblacion_censo_2010,localidades_administradas,latitud,longitud,sitio_web,fuente)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (code, row.get('provincia'), row.get('departamento'), row.get('nombre_del_gobierno_local'),
             row.get('categoria_de_gobierno'), pobl2010, row.get('localidades_administradas_por_el_gobierno_local'),
             lat, lon, row.get('sitio_web'), 'refeglo')
        )
        entity_id = f'gobierno_local:{code}'
        cur = conn.execute('SELECT 1 FROM geo_entidad WHERE id_entidad=?', (entity_id,)).fetchone()
        if cur:
            conn.execute(
                '''UPDATE geo_entidad
                   SET codigo_refeglo=?, centroid_lat=COALESCE(centroid_lat,?), centroid_lon=COALESCE(centroid_lon,?),
                       fuente_geografica=COALESCE(fuente_geografica,'refeglo')
                   WHERE id_entidad=?''',
                (code, lat, lon, entity_id)
            )
        else:
            name = row.get('nombre_del_gobierno_local') or code
            conn.execute(
                '''INSERT INTO geo_entidad
                   (id_entidad,tipo_entidad,codigo_refeglo,nombre,nombre_normalizado,
                    fuente_geografica,clasificacion_fuente,metodo_geometria,nivel_confianza,anio_fuente,
                    centroid_lat,centroid_lon,atributos_json)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (entity_id, 'gobierno_local', code, name, normalize_text(name),
                 'refeglo', 'oficial_directa', 'centroide_refeglo', 'alta', 2023, lat, lon, '{}')
            )


def refresh_control(conn: sqlite3.Connection) -> None:
    conn.execute('DELETE FROM control_cobertura')
    rows = conn.execute(
        '''SELECT g.tipo_entidad,
                  COUNT(*) entidades,
                  SUM(CASE WHEN g.centroid_lat IS NOT NULL AND g.centroid_lon IS NOT NULL THEN 1 ELSE 0 END) con_centroides,
                  SUM(CASE WHEN g.geom_wkt IS NOT NULL THEN 1 ELSE 0 END) con_geom_wkt,
                  SUM(CASE WHEN p.id_poblacion IS NOT NULL THEN 1 ELSE 0 END) con_poblacion_2022
           FROM geo_entidad g
           LEFT JOIN censo_poblacion p ON p.id_entidad = g.id_entidad AND p.anio_censo = 2022
           GROUP BY g.tipo_entidad'''
    ).fetchall()
    conn.executemany('INSERT INTO control_cobertura VALUES (?,?,?,?,?,?)', [(*r, 'Semilla; polígonos completos se generan con load_to_gpkg/build_postgis.') for r in rows])


def main() -> None:
    parser = argparse.ArgumentParser(description='Crea SQLite semilla con fuentes verificadas descargadas en el paquete.')
    parser.add_argument('--seed-dir', required=True)
    parser.add_argument('--out', required=True)
    args = parser.parse_args()
    seed_dir = Path(args.seed_dir)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists():
        out.unlink()
    conn = sqlite_connect(out)
    init_schema(conn)
    insert_sources(conn)
    load_georef_provinces(seed_dir, conn)
    load_censo_resumen(seed_dir, conn)
    load_censo_gobiernos(seed_dir, conn)
    load_refeglo(seed_dir, conn)
    refresh_control(conn)
    conn.commit()
    print(f'SQLite semilla creado: {out}')


if __name__ == '__main__':
    main()
